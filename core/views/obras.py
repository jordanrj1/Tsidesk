import io, zipfile, os, re
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q
from ..models import (Obra, Contrato, Documento, CierreMensual,
                      BodegaObra, CatalogoMaterial, TipoDocumento, HistorialMaterial,
                      Trabajador, Especialidad, DocumentoGenerado, Traslado,
                      ContratoHistorial,
                      get_checklist_trabajador, get_checklist_contrato)


def _historial(contrato, estado_nuevo, usuario, descripcion=''):
    if contrato.estado != estado_nuevo:
        ContratoHistorial.objects.create(
            contrato=contrato,
            estado_anterior=contrato.estado,
            estado_nuevo=estado_nuevo,
            descripcion=descripcion,
            usuario=usuario,
        )
from ..forms import ObraForm


@login_required
def obras_list(request):
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    qs = Obra.objects.filter(activo=True, archivada=False)
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(constructora_mandante__icontains=q))
    if estado:
        qs = qs.filter(estado=estado)
    archivadas = Obra.objects.filter(activo=True, archivada=True).order_by('nombre')
    context = {'obras': qs, 'q': q, 'estado': estado, 'archivadas': archivadas}
    return render(request, 'obras/list.html', context)


@login_required
def obra_archivar(request, pk):
    if request.method == 'POST':
        obra = get_object_or_404(Obra, pk=pk, activo=True)
        if not obra.archivada and obra.estado != 'Cerrada':
            messages.error(request, 'Solo se puede archivar una obra en estado "En cierre".')
            return redirect('obra_detail', pk=pk)
        obra.archivada = not obra.archivada
        obra.save()
        if obra.archivada:
            messages.success(request, f'Obra "{obra.nombre}" archivada. Los documentos quedan en solo lectura.')
        else:
            messages.success(request, f'Obra "{obra.nombre}" desarchivada.')
    return redirect('obras_list')


@login_required
def obra_check_trabajador(request, pk):
    """AJAX: returns vigente contracts in other obras + finiquito-pending in this obra."""
    rut = request.GET.get('rut', '').strip()
    if not rut:
        return JsonResponse({'vigente_otras_obras': [], 'finiquito_pendiente_misma_obra': []})
    contratos_vigentes = Contrato.objects.filter(
        trabajador_id=rut, estado='Vigente', activo=True
    ).exclude(obra_id=pk).select_related('obra')
    data = [
        {'obra_nombre': c.obra.nombre, 'obra_pk': c.obra.pk, 'contrato_pk': c.pk}
        for c in contratos_vigentes
    ]
    # Contratos anteriores finalizados — en ESTA obra y en OTRAS
    ESTADOS_TERMINO = ('Finalizado', 'Finiquitado', 'Rescindido', 'Trasladado')
    todos_anteriores = Contrato.objects.filter(
        trabajador_id=rut, estado__in=ESTADOS_TERMINO, activo=True,
    ).select_related('obra')

    pendientes_misma = []
    pendientes_otras = []
    for c in todos_anteriores:
        tiene_finiquito = Documento.objects.filter(
            contrato=c, tipo_documento__nombre='Finiquito Legalizado', activo=True,
        ).exists()
        if not tiene_finiquito:
            entry = {
                'contrato_pk': c.pk,
                'estado': c.estado,
                'fecha_termino': str(c.fecha_termino_real) if c.fecha_termino_real else None,
            }
            if c.obra_id == int(pk):
                pendientes_misma.append(entry)
            else:
                entry['obra_nombre'] = c.obra.nombre
                pendientes_otras.append(entry)

    return JsonResponse({
        'vigente_otras_obras': data,
        'finiquito_pendiente_misma_obra': pendientes_misma,
        'finiquito_pendiente_otras_obras': pendientes_otras,
    })


@login_required
def obra_asignar_trabajador(request, pk):
    """Agrega un trabajador a la obra. Todos los campos del contrato son opcionales.
    Estado resultante:
      - archivo subido + datos completos → Vigente
      - datos completos sin archivo      → Pendiente de Firma
      - datos incompletos                → Borrador
    """
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    from django.urls import reverse
    _redir = reverse('obra_detail', args=[pk]) + '?tab=dotacion'

    if obra.archivada or obra.estado == 'Cerrada':
        messages.error(request, f'No se pueden agregar trabajadores a una obra {("archivada" if obra.archivada else "cerrada")}.')
        return redirect(_redir)

    rut = request.POST.get('trabajador', '').strip()
    ESTADOS_ACTIVOS = ('Borrador', 'Pendiente de Firma', 'Vigente', 'En Licencia', 'Reactivado')
    ESTADOS_TERMINO = ('Finalizado', 'Finiquitado', 'Rescindido', 'Trasladado')

    try:
        trabajador = Trabajador.objects.get(rut=rut, activo=True)

        if trabajador.en_lista_negra:
            messages.error(request,
                f'{trabajador.nombre_completo} está en lista negra y no puede ser asignado. '
                f'Motivo: {trabajador.motivo_lista_negra or "sin detalle"}.')
            return redirect(_redir)

        contrato_existente = Contrato.objects.filter(
            trabajador=trabajador, obra=obra, estado__in=ESTADOS_ACTIVOS, activo=True
        ).first()
        if contrato_existente:
            messages.warning(request,
                f'{trabajador.nombre_completo} ya tiene un contrato activo en esta obra '
                f'(estado: {contrato_existente.estado}).')
            return redirect(_redir)

        # Campos opcionales
        import decimal
        from datetime import date as _date

        especialidad_id = request.POST.get('especialidad', '').strip()
        tipo_contrato   = request.POST.get('tipo_contrato', 'Plazo Fijo')
        sueldo_raw      = request.POST.get('sueldo_base', '').strip()
        fecha_inicio_s  = request.POST.get('fecha_inicio', '').strip()
        fecha_termino_s = request.POST.get('fecha_termino_estimada', '').strip()
        archivo         = request.FILES.get('archivo_contrato')

        especialidad   = Especialidad.objects.filter(pk=int(especialidad_id), activo=True).first() if especialidad_id else None
        sueldo         = None
        fecha_inicio   = None
        fecha_termino  = None

        if sueldo_raw:
            try:
                sueldo = decimal.Decimal(sueldo_raw.replace('.', '').replace(',', '.')).quantize(
                    decimal.Decimal('1'), rounding=decimal.ROUND_HALF_UP)
            except decimal.InvalidOperation:
                pass

        if fecha_inicio_s:
            try:
                fecha_inicio = _date.fromisoformat(fecha_inicio_s)
            except ValueError:
                pass

        if fecha_termino_s:
            try:
                fecha_termino = _date.fromisoformat(fecha_termino_s)
            except ValueError:
                pass

        if fecha_inicio and fecha_termino and fecha_termino < fecha_inicio:
            messages.error(request, 'La fecha de término no puede ser anterior a la fecha de inicio.')
            return redirect(_redir)

        # Determinar estado
        datos_completos = bool(especialidad and sueldo and fecha_inicio)
        if archivo and datos_completos:
            estado = 'Vigente'
        elif datos_completos:
            estado = 'Pendiente de Firma'
        else:
            estado = 'Borrador'

        contrato_previo = Contrato.objects.filter(
            trabajador=trabajador, obra=obra, estado__in=ESTADOS_TERMINO, activo=True
        ).order_by('-creado_el').first()

        contrato = Contrato.objects.create(
            trabajador=trabajador,
            obra=obra,
            especialidad=especialidad,
            tipo_contrato=tipo_contrato,
            sueldo_base=sueldo,
            fecha_inicio=fecha_inicio,
            fecha_termino_estimada=fecha_termino,
            estado=estado,
            es_recontratacion=bool(contrato_previo),
            contrato_anterior=contrato_previo,
        )

        # Subir archivo si llegó
        if archivo:
            from ..models import TipoDocumento as _TD
            tipo_doc = _TD.objects.filter(activo=True, nivel='Contrato', nombre__icontains='contrato').first()
            if tipo_doc:
                Documento.objects.create(
                    tipo_documento=tipo_doc,
                    contrato=contrato,
                    archivo=archivo,
                    usuario_carga=request.user.username,
                )

        # Mensaje de resultado
        estado_label = {'Vigente': 'Vigente', 'Pendiente de Firma': 'Pendiente de Firma', 'Borrador': 'Borrador — completa los datos cuando llegue el contrato firmado'}[estado]
        if contrato_previo:
            tiene_finiquito = Documento.objects.filter(
                contrato=contrato_previo, tipo_documento__nombre='Finiquito Legalizado', activo=True
            ).exists()
            if not tiene_finiquito:
                messages.warning(request,
                    f'Re-contratación de {trabajador.nombre_completo} registrada ({estado_label}). '
                    f'El contrato anterior no tiene finiquito subido.')
            else:
                messages.success(request, f'{trabajador.nombre_completo} re-contratado/a ({estado_label}).')
        else:
            messages.success(request, f'{trabajador.nombre_completo} agregado a la obra ({estado_label}).')

    except Trabajador.DoesNotExist:
        messages.error(request, 'Trabajador no encontrado.')
    except Exception as e:
        messages.error(request, f'Error al agregar trabajador: {e}')

    return redirect(_redir)


@login_required
def obra_ingreso_rapido(request, pk):
    """Alias mantenido por compatibilidad — redirige a obra_asignar_trabajador."""
    return obra_asignar_trabajador(request, pk)


@login_required
def obra_completar_borrador(request, pk, contrato_pk):
    """Completa los datos de un Contrato en estado Borrador."""
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, estado='Borrador', activo=True)
    especialidades = Especialidad.objects.filter(activo=True).order_by('nombre')
    from ..models import TipoDocumento

    if request.method == 'POST':
        especialidad_id = request.POST.get('especialidad', '').strip()
        tipo_contrato = request.POST.get('tipo_contrato', 'Plazo Fijo')
        sueldo_raw = request.POST.get('sueldo_base', '').strip()
        fecha_inicio = request.POST.get('fecha_inicio', '').strip() or None
        fecha_termino = request.POST.get('fecha_termino_estimada', '').strip() or None
        archivo = request.FILES.get('archivo_contrato')

        try:
            import decimal
            from datetime import date as _date

            if fecha_inicio and fecha_termino:
                fi = _date.fromisoformat(fecha_inicio)
                ft = _date.fromisoformat(fecha_termino)
                if ft < fi:
                    messages.error(request, 'La fecha de término no puede ser anterior a la fecha de inicio.')
                    return redirect(request.path)

            contrato.tipo_contrato = tipo_contrato
            if especialidad_id:
                contrato.especialidad = Especialidad.objects.get(pk=int(especialidad_id), activo=True)
            if sueldo_raw:
                contrato.sueldo_base = decimal.Decimal(sueldo_raw.replace(',', '')).quantize(decimal.Decimal('1'), rounding=decimal.ROUND_HALF_UP)
            if fecha_inicio:
                contrato.fecha_inicio = fecha_inicio
            if fecha_termino:
                contrato.fecha_termino_estimada = fecha_termino

            if archivo:
                tipo_contrato_doc = TipoDocumento.objects.filter(
                    activo=True, nivel='Contrato', nombre__icontains='contrato'
                ).first()
                if tipo_contrato_doc:
                    from ..models import Documento
                    Documento.objects.create(
                        tipo_documento=tipo_contrato_doc,
                        contrato=contrato,
                        archivo=archivo,
                        usuario_carga=request.user.username,
                    )
                contrato.estado = 'Vigente'
                messages.success(request, f'Datos completados. Contrato de {contrato.trabajador.nombre_completo} ahora está Vigente.')
            else:
                contrato.estado = 'Pendiente de Firma'
                messages.success(request, f'Datos completados. Contrato de {contrato.trabajador.nombre_completo} está Pendiente de Firma.')

            contrato.save()
            from django.urls import reverse
            return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')

        except Exception as e:
            messages.error(request, f'Error al guardar: {e}')

    tipos_contrato = Contrato.TIPO_CONTRATO_CHOICES
    return render(request, 'obras/completar_borrador.html', {
        'obra': obra,
        'contrato': contrato,
        'especialidades': especialidades,
        'tipos_contrato': tipos_contrato,
    })


@login_required
def obra_quitar_trabajador(request, pk, contrato_pk):
    """Gestiona la baja de un trabajador de una obra con diálogo de decisión."""
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)

    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        if accion == 'eliminar_erroneo':
            # Ingreso erróneo: desactivar contrato sin rastro en planilla
            _historial(contrato, 'Rescindido', request.user.username, 'Eliminado: ingreso erróneo')
            contrato.activo = False
            contrato.estado = 'Rescindido'
            contrato.save()
            # Desactivar DocumentoGenerado vinculado si existe y estaba Pendiente
            DocumentoGenerado.objects.filter(contrato=contrato, activo=True).update(activo=False)
            messages.success(request, f'{contrato.trabajador.nombre_completo} eliminado de la obra (ingreso erróneo).')

        elif accion == 'mantener':
            messages.info(request, 'Se mantuvo el registro del trabajador.')

        elif accion == 'actualizar_contrato':
            from django.urls import reverse
            return redirect(
                reverse('contrato_edit', args=[contrato.pk])
                + f'?next={reverse("obra_detail", args=[pk])}%3Ftab%3Ddotacion'
            )

        elif accion == 'eliminar_doc_contrato':
            DocumentoGenerado.objects.filter(
                contrato=contrato, tipo='contrato_trabajo', activo=True
            ).update(activo=False)
            messages.success(request, 'Documentación del contrato eliminada.')

        from django.urls import reverse
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')

    # GET: mostrar modal de decisión (manejado desde template vía AJAX/modal)
    from django.urls import reverse
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_licencia_trabajador(request, pk, contrato_pk):
    """Poner un trabajador en estado Licencia Médica."""
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)
    fecha_inicio_lic = request.POST.get('fecha_inicio_licencia', '').strip()
    obs = request.POST.get('obs_licencia', '').strip()
    if not fecha_inicio_lic:
        messages.error(request, 'Debe indicar la fecha de inicio de la licencia.')
    else:
        _historial(contrato, 'En Licencia', request.user.username, f'Licencia desde {fecha_inicio_lic}. {obs}'.strip())
        contrato.estado = 'En Licencia'
        contrato.fecha_inicio_licencia = fecha_inicio_lic
        contrato.fecha_fin_licencia = None
        contrato.obs_licencia = obs
        contrato.save()
        messages.success(request, f'{contrato.trabajador.nombre_completo} puesto en Licencia Médica desde {fecha_inicio_lic}.')
    from django.urls import reverse
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_reactivar_trabajador(request, pk, contrato_pk):
    """Reactivar un trabajador post-licencia y generar Acta de Reactivación."""
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)
    fecha_retorno = request.POST.get('fecha_retorno', '').strip()
    obs = request.POST.get('obs_reactivacion', '').strip()
    if not fecha_retorno:
        messages.error(request, 'Debe indicar la fecha de retorno.')
        from django.urls import reverse
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')
    # Validar que fecha retorno >= fecha inicio licencia
    if contrato.fecha_inicio_licencia:
        from datetime import date as _date
        try:
            if _date.fromisoformat(fecha_retorno) < contrato.fecha_inicio_licencia:
                messages.error(request, 'La fecha de retorno no puede ser anterior a la fecha de inicio de licencia.')
                from django.urls import reverse
                return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')
        except ValueError:
            pass

    _historial(contrato, 'Reactivado', request.user.username, f'Retorno {fecha_retorno}. {obs}'.strip())
    contrato.estado = 'Reactivado'
    contrato.fecha_fin_licencia = fecha_retorno
    contrato.obs_licencia = (contrato.obs_licencia + '\n' + obs).strip() if obs else contrato.obs_licencia
    contrato.save()

    # Generar Acta de Reactivación
    empresa = obra.empresa or None
    if empresa:
        doc_acta = DocumentoGenerado.objects.create(
            tipo='acta_reactivacion',
            empresa=empresa,
            trabajador=contrato.trabajador,
            contrato=contrato,
            obra=obra,
            datos={
                'fecha_retorno': fecha_retorno,
                'fecha_inicio_licencia': str(contrato.fecha_inicio_licencia) if contrato.fecha_inicio_licencia else '',
                'obs_reactivacion': obs,
            },
            usuario=request.user.username,
        )
        messages.success(
            request,
            f'{contrato.trabajador.nombre_completo} reactivado. '
            f'<a href="/documentos-empresa/preview/{doc_acta.pk}/" target="_blank">Ver Acta de Reactivación</a>',
        )
    else:
        messages.success(request, f'{contrato.trabajador.nombre_completo} reactivado. (Sin empresa asignada a la obra, no se generó acta.)')

    from django.urls import reverse
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_confirmar_reactivacion(request, pk, contrato_pk):
    """Confirmar reactivación: pasa de Reactivado → Vigente."""
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)
    _historial(contrato, 'Vigente', request.user.username, 'Confirmación post-reactivación')
    contrato.estado = 'Vigente'
    contrato.save()
    messages.success(request, f'{contrato.trabajador.nombre_completo} confirmado como Vigente.')
    from django.urls import reverse
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_finalizar_contrato(request, pk, contrato_pk):
    """Finaliza un contrato Vigente/Reactivado/En Licencia → Finalizado.
    Este paso es previo a generar el finiquito. El finiquito no es obligatorio."""
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)

    ESTADOS_FINALIZABLES = ('Vigente', 'Reactivado', 'En Licencia', 'Pendiente de Firma')
    if contrato.estado not in ESTADOS_FINALIZABLES:
        messages.error(request, f'El contrato está en estado "{contrato.estado}" y no puede finalizarse desde aquí.')
        from django.urls import reverse
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')

    tipo_termino = request.POST.get('tipo_termino', '').strip()
    motivo_termino = request.POST.get('motivo_termino', '').strip()
    fecha_termino_real_str = request.POST.get('fecha_termino_real', '').strip()

    from datetime import date as _date
    fecha_termino_real = None
    if fecha_termino_real_str:
        try:
            fecha_termino_real = _date.fromisoformat(fecha_termino_real_str)
        except ValueError:
            messages.error(request, 'Fecha de término inválida.')
            from django.urls import reverse
            return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')

    _historial(
        contrato, 'Finalizado', request.user.username,
        f'Contrato finalizado. Tipo: {tipo_termino or "sin especificar"}. {motivo_termino}'.strip()
    )
    contrato.estado = 'Finalizado'
    contrato.tipo_termino = tipo_termino
    contrato.motivo_termino = motivo_termino
    if fecha_termino_real:
        contrato.fecha_termino_real = fecha_termino_real
    contrato.save()

    from django.urls import reverse
    # Si viene del modal de vencido con opción finiquito, redirigir directo al generador
    ir_a_finiquito = request.POST.get('ir_a_finiquito', '').strip()
    if ir_a_finiquito:
        messages.success(
            request,
            f'{contrato.trabajador.nombre_completo} finalizado por vencimiento de plazo. '
            f'Completa el finiquito a continuación.'
        )
        return redirect(
            reverse('doc_generado_create') +
            f'?tipo=finiquito&contrato_id={contrato.pk}&trabajador_rut={contrato.trabajador_id}'
        )

    messages.success(
        request,
        f'{contrato.trabajador.nombre_completo} finalizado. '
        f'El finiquito puede subirse en la Carpeta de la Obra → sección Finiquitos Pendientes.'
    )
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_resolver_duplicado(request, pk, contrato_pk):
    """Elimina el contrato duplicado inferior (Pendiente de Firma) manteniendo el de mayor jerarquía."""
    if request.method != 'POST':
        return redirect('obra_detail', pk=pk)
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)
    contrato.activo = False
    contrato.estado = 'Rescindido'
    contrato.save()
    DocumentoGenerado.objects.filter(contrato=contrato, activo=True).update(activo=False)
    messages.success(request, f'Contrato duplicado #{contrato_pk} eliminado.')
    from django.urls import reverse
    return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')


@login_required
def obra_create(request):
    if request.method == 'POST':
        form = ObraForm(request.POST)
        if form.is_valid():
            obra = form.save()
            # Inicializar bodega para todos los materiales activos
            for mat in CatalogoMaterial.objects.filter(activo=True):
                BodegaObra.objects.get_or_create(obra=obra, material=mat, defaults={'stock_actual': 0})
            messages.success(request, f'Obra "{obra.nombre}" creada exitosamente.')
            return redirect('obra_detail', pk=obra.pk)
    else:
        form = ObraForm()
    return render(request, 'obras/form.html', {'form': form, 'titulo': 'Nueva Obra'})


@login_required
def obra_edit(request, pk):
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    if request.method == 'POST':
        form = ObraForm(request.POST, instance=obra)
        if form.is_valid():
            form.save()
            messages.success(request, 'Obra actualizada correctamente.')
            return redirect('obra_detail', pk=pk)
    else:
        form = ObraForm(instance=obra)
    return render(request, 'obras/form.html', {'form': form, 'titulo': 'Editar Obra', 'obra': obra})


@login_required
def obra_detail(request, pk):
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    tab = request.GET.get('tab', 'resumen')
    contratos = obra.contratos.filter(activo=True).select_related('trabajador', 'especialidad').order_by('-creado_el')
    documentos = Documento.objects.filter(obra=obra, activo=True).select_related('tipo_documento')
    bodega = obra.bodega_items.all().select_related('material').order_by('material__nombre')
    from ..models import HistorialMaterial
    historial_bodega = HistorialMaterial.objects.filter(
        obra=obra, tipo_movimiento='ENTREGA_TERRENO'
    ).select_related('material', 'trabajador_capataz').order_by('-fecha_movimiento')
    cierres = obra.cierres_mensuales.all().order_by('-anio', '-mes')

    # Documentos de trabajadores activos en la obra — agrupados por trabajador
    _docs_qs = Documento.objects.filter(
        Q(trabajador_rut__in=contratos.filter(estado='Vigente').values_list('trabajador_id', flat=True)) |
        Q(contrato__in=contratos),
        activo=True
    ).select_related('tipo_documento', 'contrato', 'contrato__trabajador').order_by(
        'trabajador_rut', 'tipo_documento__nombre'
    )

    # Build grouped structure: {rut: {nombre, rut, docs: []}}
    _grupos = {}
    for doc in _docs_qs:
        if doc.contrato and doc.contrato.trabajador:
            trab = doc.contrato.trabajador
            key = trab.rut
            if key not in _grupos:
                _grupos[key] = {'nombre': trab.nombre_completo, 'rut': trab.rut, 'docs': []}
        elif doc.trabajador_rut:
            key = doc.trabajador_rut
            if key not in _grupos:
                _grupos[key] = {'nombre': doc.trabajador_rut, 'rut': doc.trabajador_rut, 'docs': []}
        else:
            continue
        _grupos[key]['docs'].append(doc)

    docs_trabajadores = list(_grupos.values())

    # Dotación: solo excluir trabajadores con contrato ACTIVO en esta obra
    # (Finalizado/Finiquitado no bloquea — pueden ser re-contratados)
    ESTADOS_ACTIVOS_OBRA = ('Borrador', 'Pendiente de Firma', 'Vigente', 'En Licencia', 'Reactivado')
    ruts_con_contrato_activo = contratos.filter(
        estado__in=ESTADOS_ACTIVOS_OBRA
    ).values_list('trabajador_id', flat=True)
    trabajadores_disponibles = Trabajador.objects.filter(activo=True).exclude(
        rut__in=ruts_con_contrato_activo
    ).order_by('nombres', 'apellidos')
    especialidades = Especialidad.objects.filter(activo=True).order_by('nombre')
    _dg_map = {
        dg.contrato_id: dg
        for dg in DocumentoGenerado.objects.filter(
            contrato__in=contratos, tipo='contrato_trabajo', activo=True
        )
    }
    # Contrato firmado subido: obtener objeto para poder previsualizar
    _doc_firmado_map = {
        doc.contrato_id: doc
        for doc in Documento.objects.filter(
            contrato__in=contratos,
            tipo_documento__nombre='Contrato de Trabajo Firmado',
            activo=True,
        ).exclude(archivo=None).exclude(archivo='')
    }
    _ESTADOS_DOTACION = {'Borrador', 'Pendiente de Firma', 'Vigente', 'En Licencia', 'Reactivado'}
    contratos_con_doc = [
        (c, _dg_map.get(c.pk), _doc_firmado_map.get(c.pk))
        for c in contratos if c.estado in _ESTADOS_DOTACION
    ]
    contratos_historial = [c for c in contratos if c.estado in ('Finalizado', 'Finiquitado')]

    # Detectar contratos duplicados: solo cuentan estados activos (no terminados)
    # Un Finalizado + Pendiente de Firma es una re-contratación válida, no un duplicado
    from collections import Counter
    _ESTADOS_ACTIVOS_DUP = {'Borrador', 'Pendiente de Firma', 'Vigente', 'En Licencia', 'Reactivado'}
    rut_count = Counter(c.trabajador_id for c in contratos if c.estado in _ESTADOS_ACTIVOS_DUP)
    contratos_duplicados = {rut for rut, n in rut_count.items() if n > 1}

    # Carpeta de Obra: filters
    q_carpeta = request.GET.get('q_carpeta', '').strip()
    tipo_carpeta = request.GET.get('tipo_carpeta', '').strip()
    documentos_carpeta = documentos  # base queryset (already filtered obra=obra, activo=True)
    if q_carpeta:
        documentos_carpeta = documentos_carpeta.filter(tipo_documento__nombre__icontains=q_carpeta)
    if tipo_carpeta:
        documentos_carpeta = documentos_carpeta.filter(tipo_documento_id=tipo_carpeta)
    tipos_docs_obra = TipoDocumento.objects.filter(nivel='Obra', activo=True).order_by('nombre')

    from datetime import timedelta as _td
    from ..models import LicenciaMedica as _Lic
    hoy_detail = timezone.now().date()

    # Contratos próximos a vencer (en los próximos 10 días) — usa fecha_extension si existe
    _ESTADOS_CON_FECHA = {'Vigente', 'Pendiente de Firma', 'En Licencia', 'Reactivado'}
    contratos_por_vencer = [
        c for c in contratos
        if c.estado in _ESTADOS_CON_FECHA and c.fecha_vigencia_efectiva
        and 0 <= (c.fecha_vigencia_efectiva - hoy_detail).days <= 10
    ]
    # Contratos vencidos sin finalizar — si tiene fecha_extension vigente, NO aparece aquí
    contratos_vencidos_sin_accion = [
        c for c in contratos
        if c.estado in _ESTADOS_CON_FECHA and c.fecha_vigencia_efectiva
        and c.fecha_vigencia_efectiva < hoy_detail
    ]
    # Finalizados sin finiquito subido (carpeta de finiquitos pendientes)
    contratos_finalizados = obra.contratos.filter(
        estado='Finalizado', activo=True
    ).select_related('trabajador')
    finalizados_sin_finiquito = [
        c for c in contratos_finalizados
        if not Documento.objects.filter(
            contrato=c,
            tipo_documento__nombre__icontains='finiquito',
            activo=True
        ).exists()
    ]
    # Días para cierre estimado de la obra
    dias_para_cierre_obra = (obra.fecha_termino_estimada - hoy_detail).days if obra.fecha_termino_estimada else None

    # IDs de contratos con licencia Tipo 2 activa (accidente laboral)
    licencias_tipo2_activas = set(
        _Lic.objects.filter(
            contrato__in=contratos, tipo='2', activo=True,
            estado__in=('Presentada', 'En trámite', 'Autorizada', 'Prorrogada')
        ).exclude(fecha_fin__lt=hoy_detail).values_list('contrato_id', flat=True)
    )

    # Carpetas completas: solo evaluar cuando está en cierre
    carpetas_completas = False
    if obra.estado == 'Cerrada' and not obra.archivada:
        from ..models import get_checklist_contrato as _get_cl
        contratos_term = contratos.filter(estado__in=['Finalizado', 'Finiquitado'])
        if contratos_term.exists():
            carpetas_completas = all(
                all(item['estado'] in ('ok', 'proximo') for item in _get_cl(c))
                for c in contratos_term
            )

    context = {
        'obra': obra,
        'contratos': contratos,
        'contratos_con_doc': contratos_con_doc,
        'contratos_duplicados': contratos_duplicados,
        'contratos_por_vencer': contratos_por_vencer,
        'contratos_vencidos_sin_accion': contratos_vencidos_sin_accion,
        'finalizados_sin_finiquito': finalizados_sin_finiquito,
        'contratos_historial': contratos_historial,
        'licencias_tipo2_activas': licencias_tipo2_activas,
        'carpetas_completas': carpetas_completas,
        'trabajadores_disponibles': trabajadores_disponibles,
        'especialidades': especialidades,
        'documentos': documentos,
        'documentos_carpeta': documentos_carpeta,
        'tipos_docs_obra': tipos_docs_obra,
        'q_carpeta': q_carpeta,
        'tipo_carpeta': tipo_carpeta,
        'docs_trabajadores': docs_trabajadores,
        'bodega': bodega,
        'historial_bodega': historial_bodega,
        'cierres': cierres,
        'tab': tab,
        'dias_para_cierre_obra': dias_para_cierre_obra,
        'TIPO_TERMINO_CHOICES': Contrato.TIPO_TERMINO_CHOICES,
    }
    return render(request, 'obras/detail.html', context)


@login_required
def obra_cierre_check(request, pk):
    from ..models import get_checklist_contrato
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    n_vigentes = obra.contratos.filter(estado='Vigente', activo=True).count()
    n_pendientes_firma = obra.contratos.filter(estado='Pendiente de Firma', activo=True).count()
    n_en_licencia = obra.contratos.filter(estado__in=['En Licencia', 'Reactivado'], activo=True).count()

    puede_cerrar = n_vigentes == 0 and n_pendientes_firma == 0 and n_en_licencia == 0

    # Carpetas completas: todos los contratos terminados tienen sus docs ok
    contratos_term = obra.contratos.filter(estado__in=['Finalizado', 'Finiquitado'], activo=True)
    carpetas_completas = False
    if contratos_term.exists() and puede_cerrar:
        carpetas_completas = all(
            all(item['estado'] in ('ok', 'proximo') for item in get_checklist_contrato(c))
            for c in contratos_term
        )

    return JsonResponse({
        'contratos_vigentes': n_vigentes,
        'contratos_pendientes_firma': n_pendientes_firma,
        'contratos_en_licencia': n_en_licencia,
        'puede_cerrar': puede_cerrar,
        'carpetas_completas': carpetas_completas,
    })


@login_required
def obra_cerrar(request, pk):
    if request.method == 'POST':
        obra = get_object_or_404(Obra, pk=pk, activo=True)
        n_vigentes = obra.contratos.filter(estado='Vigente', activo=True).count()
        n_pendientes_firma = obra.contratos.filter(estado='Pendiente de Firma', activo=True).count()
        n_en_licencia = obra.contratos.filter(estado__in=['En Licencia', 'Reactivado'], activo=True).count()

        bloqueos = []
        if n_vigentes:
            bloqueos.append(f'{n_vigentes} contrato(s) aún vigente(s)')
        if n_pendientes_firma:
            bloqueos.append(f'{n_pendientes_firma} contrato(s) pendiente(s) de firma')
        if n_en_licencia:
            bloqueos.append(f'{n_en_licencia} trabajador(es) en licencia o reactivado')

        if bloqueos:
            messages.error(request, f'No se puede cerrar la obra: {"; ".join(bloqueos)}.')
            from django.urls import reverse as _rev
            return redirect(_rev('obra_detail', args=[pk]) + '?tab=dotacion')

        obra.estado = 'Cerrada'
        obra.fecha_termino_real = timezone.now().date()
        obra.save()
        messages.success(
            request,
            f'Obra "{obra.nombre}" puesta en cierre. '
            f'Puedes seguir subiendo documentación hasta completar todas las carpetas.'
        )
    return redirect('obras_list')


@login_required
def obra_download_zip(request, pk):
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    docs = Documento.objects.filter(
        Q(obra=obra) |
        Q(contrato__obra=obra) |
        Q(trabajador_rut__in=Contrato.objects.filter(obra=obra).values_list('trabajador_id', flat=True)),
        activo=True
    ).select_related('tipo_documento')

    cierres = CierreMensual.objects.filter(obra=obra).order_by('anio', 'mes', 'fecha_cierre')

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for doc in docs:
            if doc.archivo and os.path.exists(doc.archivo.path):
                tipo_nombre = doc.tipo_documento.nombre if doc.tipo_documento else 'Documento'
                fname = os.path.basename(doc.archivo.path)
                if doc.contrato:
                    rut = doc.contrato.trabajador.rut
                    arcname = f"trabajadores/{rut}/{tipo_nombre}_{fname}"
                elif doc.trabajador_rut:
                    arcname = f"trabajadores/{doc.trabajador_rut}/{tipo_nombre}_{fname}"
                else:
                    arcname = f"carpeta_obra/{tipo_nombre}_{fname}"
                zf.write(doc.archivo.path, arcname)

        for cierre in cierres:
            if cierre.archivo_consolidado:
                try:
                    path = cierre.archivo_consolidado.path
                    if os.path.exists(path):
                        arcname = f"cierres/{os.path.basename(path)}"
                        zf.write(path, arcname)
                except Exception:
                    pass

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    nombre_obra = obra.nombre.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="obra_{nombre_obra}.zip"'
    return response


def _sheet_remuneraciones(ws, obra, fecha_inicio, fecha_fin, filas, descripcion=''):
    """Escribe la hoja de planilla de remuneraciones en un worksheet dado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    ws.title = 'Remuneraciones'

    headers = [
        'N°', 'Nombre Completo', 'RUT', 'Especialidad',
        'Tipo Contrato', 'Días\nPeríodo', 'Días\nLicencia', 'Días\nTrabajados',
        'Sueldo Base ($)', 'Sueldo\nProp. ($)',
        'AFP ($)', 'Salud ($)', 'Cesantía ($)', 'Total Desc. ($)', 'Líquido ($)',
    ]
    NCOLS = len(headers)
    last_col = openpyxl.utils.get_column_letter(NCOLS)

    ws.merge_cells(f'A1:{last_col}1')
    titulo = ws['A1']
    titulo.value = f'PLANILLA DE REMUNERACIONES — {obra.nombre}'
    titulo.font = Font(bold=True, size=13, color='FFFFFF')
    titulo.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last_col}2')
    periodo = ws['A2']
    desc_txt = f'  |  {descripcion}' if descripcion else ''
    periodo.value = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}  |  Generado el {date.today().strftime("%d/%m/%Y")}{desc_txt}'
    periodo.font = Font(size=10, italic=True, color='555555')
    periodo.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 36

    fill_par = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_licencia = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    num_font = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    for i, f in enumerate(filas, 1):
        row = i + 3
        tiene_licencia = f.get('dias_licencia', 0) > 0
        relleno = fill_licencia if tiene_licencia else (fill_par if i % 2 == 0 else fill_impar)
        data = [
            i,
            f['nombre'],
            f['rut'],
            f['especialidad'],
            f['tipo_contrato'],
            f.get('dias', 0),
            f.get('dias_licencia', 0),
            f.get('dias_trabajados', f.get('dias', 0)),
            f['sueldo_base'],
            f.get('sueldo_proporcional', f['sueldo_base']),
            f.get('desc_afp', ''),
            f.get('desc_salud', ''),
            f.get('desc_cesantia', ''),
            f.get('total_descuentos', ''),
            f.get('liquido', ''),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = num_font
            cell.fill = relleno
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 6, 7, 8) else left
            if col >= 9 and val != '':
                cell.number_format = '#,##0'

    total_row = len(filas) + 4
    ws.cell(row=total_row, column=2, value=f'TOTAL TRABAJADORES: {len(filas)}').font = Font(bold=True, size=10)
    for col in range(1, NCOLS + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        c.border = border

    col_widths = [5, 30, 15, 22, 16, 10, 10, 11, 16, 16, 12, 12, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'


def _sheet_checklist(ws, obra, fecha_inicio, fecha_fin, workers, tipos_personal,
                     tipos_contrato, tipos_obra, obra_docs, sheet_num, total_sheets):
    """Escribe la hoja de checklist documental formateada para impresión carta/A4 apaisado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    suffix = f' ({sheet_num}/{total_sheets})' if total_sheets > 1 else ''
    ws.title = f'Checklist{suffix}'

    # Page setup: A4 landscape, fit to 1 page wide
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Styles
    thin = Side(style='thin', color='CCCCCC')
    medium = Side(style='medium', color='999999')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_vencido = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_pendiente = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_header_dark = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    fill_header_blue = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    fill_header_orange = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
    fill_section = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    fill_alt = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

    font_white_bold = Font(bold=True, color='FFFFFF', size=11)
    font_white_sm = Font(bold=True, color='FFFFFF', size=9)
    font_ok = Font(bold=True, color='006100', size=11)
    font_vencido = Font(bold=True, color='9C0006', size=11)
    font_pendiente = Font(bold=True, color='7F4F00', size=11)
    font_normal = Font(size=9)
    font_bold_sm = Font(bold=True, size=9)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    rotated = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Columnas fijas: N°, Trabajador, RUT, Especialidad
    FIXED_COLS = 4

    def write_title(row, text, fill, span):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.font = font_white_bold
        c.fill = fill
        c.alignment = center
        ws.row_dimensions[row].height = 22

    def write_section_header(row, text, fill, span):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill
        c.alignment = left
        ws.row_dimensions[row].height = 18

    current_row = 1

    # ── Título principal ──────────────────────────────────────────────
    total_cols_personal = FIXED_COLS + len(tipos_personal)
    total_cols_contrato = FIXED_COLS + len(tipos_contrato)
    max_cols = max(total_cols_personal, total_cols_contrato, FIXED_COLS + len(tipos_obra) + 1, 6)

    write_title(current_row, f'CHECKLIST DOCUMENTAL — OBRA: {obra.nombre}', fill_header_dark, max_cols)
    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
    sub = ws.cell(row=current_row, column=1,
                  value=f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}  |  Generado: {date.today().strftime("%d/%m/%Y")}  |  Trabajadores: {len(workers)}')
    sub.font = Font(size=9, italic=True, color='555555')
    sub.alignment = center
    ws.row_dimensions[current_row].height = 14
    current_row += 1

    def write_matrix_section(start_row, label, fill_section_color, tipos, get_estado):
        """Escribe tabla matriz de working × tipos en el worksheet. Retorna la siguiente fila disponible."""
        row = start_row

        # Separador
        row += 1
        ws.row_dimensions[row].height = 8

        # Sección header
        row += 1
        total = FIXED_COLS + len(tipos)
        write_section_header(row, f'  {label}', fill_section_color, total)
        row += 1

        if not tipos:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=FIXED_COLS + 1)
            c = ws.cell(row=row, column=1, value='Sin tipos de documentos configurados')
            c.font = Font(italic=True, color='888888', size=9)
            c.alignment = center
            row += 1
            return row

        # Header de columnas
        fixed_labels = ['N°', 'Trabajador', 'RUT', 'Especialidad']
        fixed_widths = [4, 24, 13, 14]
        for col, (lbl, w) in enumerate(zip(fixed_labels, fixed_widths), 1):
            c = ws.cell(row=row, column=col, value=lbl)
            c.font = font_white_sm
            c.fill = fill_header_dark
            c.alignment = center
            c.border = border_thin
            ws.column_dimensions[get_column_letter(col)].width = w

        for j, tipo in enumerate(tipos):
            col = FIXED_COLS + j + 1
            c = ws.cell(row=row, column=col, value=tipo.nombre)
            c.font = font_white_sm
            c.fill = fill_section_color
            c.alignment = rotated
            c.border = border_thin
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.row_dimensions[row].height = 30
        row += 1

        # Filas de trabajadores
        for i, w in enumerate(workers):
            fill_row = fill_alt if i % 2 == 0 else PatternFill(fill_type=None)
            for col, val in enumerate([i + 1, w['trabajador'].nombre_completo, w['trabajador'].rut, w['especialidad']], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = font_normal
                c.fill = fill_row
                c.border = border_thin
                c.alignment = center if col in (1, 3) else left
            for j, tipo in enumerate(tipos):
                col = FIXED_COLS + j + 1
                estado = get_estado(w, tipo.pk)
                if estado in ('ok', 'proximo'):
                    mark, fill, fnt = '✓', fill_ok, font_ok
                elif estado == 'vencido':
                    mark, fill, fnt = '✗', fill_vencido, font_vencido
                else:
                    mark, fill, fnt = '✗', fill_pendiente, font_pendiente
                c = ws.cell(row=row, column=col, value=mark)
                c.font = fnt
                c.fill = fill
                c.border = border_thin
                c.alignment = center
            ws.row_dimensions[row].height = 14
            row += 1

        return row

    # ── Sección A: Documentos Personales ─────────────────────────────
    current_row = write_matrix_section(
        current_row,
        'A) DOCUMENTOS PERSONALES (nivel trabajador)',
        fill_header_blue,
        tipos_personal,
        lambda w, pk: w['personal'].get(pk, 'pendiente'),
    )

    # ── Sección B: Documentos de Contrato ─────────────────────────────
    current_row = write_matrix_section(
        current_row,
        'B) DOCUMENTOS DE CONTRATO/OBRA (por trabajador)',
        fill_header_orange,
        tipos_contrato,
        lambda w, pk: w['contractual'].get(pk, 'pendiente'),
    )

    # ── Sección C: Documentos de Obra ─────────────────────────────────
    current_row += 1
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    obra_span = max_cols
    write_section_header(current_row, '  C) DOCUMENTOS DE OBRA (generales)', fill_header_dark, obra_span)
    current_row += 1

    if tipos_obra:
        for col, lbl in enumerate(['Documento', 'Estado', 'Observación'], 1):
            c = ws.cell(row=current_row, column=col, value=lbl)
            c.font = font_white_sm
            c.fill = fill_header_dark
            c.alignment = center
            c.border = border_thin
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 22
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for i, tipo in enumerate(tipos_obra):
            estado = obra_docs.get(tipo.pk, 'pendiente')
            if estado in ('ok', 'proximo'):
                estado_label, fill, fnt = 'CARGADO ✓', fill_ok, font_ok
                obs = 'Vigente' if estado == 'ok' else 'Próximo a vencer'
            elif estado == 'vencido':
                estado_label, fill, fnt = 'VENCIDO ✗', fill_vencido, font_vencido
                obs = 'Requiere renovación'
            else:
                estado_label, fill, fnt = 'FALTANTE ✗', fill_pendiente, font_pendiente
                obs = 'Pendiente de carga'

            fill_row = fill_alt if i % 2 == 0 else PatternFill(fill_type=None)
            c = ws.cell(row=current_row, column=1, value=tipo.nombre)
            c.font = Font(bold=True, size=9)
            c.fill = fill_row
            c.border = border_thin
            c.alignment = left

            c2 = ws.cell(row=current_row, column=2, value=estado_label)
            c2.font = fnt
            c2.fill = fill
            c2.border = border_thin
            c2.alignment = center

            c3 = ws.cell(row=current_row, column=3, value=obs)
            c3.font = font_normal
            c3.fill = fill_row
            c3.border = border_thin
            c3.alignment = left

            ws.row_dimensions[current_row].height = 14
            current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        c = ws.cell(row=current_row, column=1, value='Sin documentos de obra configurados')
        c.font = Font(italic=True, color='888888', size=9)
        c.alignment = center

    # Leyenda
    current_row += 2
    ws.cell(row=current_row, column=1, value='LEYENDA:').font = Font(bold=True, size=8)
    for col, (sym, txt, fill) in enumerate([('✓', 'Cargado/Vigente', fill_ok), ('✗', 'Faltante/Vencido', fill_vencido), ('✗', 'Sin cargar', fill_pendiente)], 2):
        c = ws.cell(row=current_row, column=col, value=f'{sym} {txt}')
        c.fill = fill
        c.font = Font(size=8, bold=True, color='000000')
        c.alignment = center
        c.border = border_thin
    ws.row_dimensions[current_row].height = 14


def _sheet_bodega(ws, obra, fecha_inicio, fecha_fin):
    """Hoja 3: todos los materiales ingresados a la obra (sin filtro de período)."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws.title = 'Materiales Ingresados'

    fill_dark = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    fill_hdr  = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    fill_alt  = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left      = Alignment(horizontal='left', vertical='center')

    # ── Encabezado ────────────────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f'MATERIALES INGRESADOS A LA OBRA — {obra.nombre}'
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = fill_dark
    c.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = f'Generado el {date.today().strftime("%d/%m/%Y")}  |  Período del cierre: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
    c.font = Font(size=9, italic=True, color='555555')
    c.alignment = center
    ws.row_dimensions[2].height = 15

    # ── Cabecera de tabla ─────────────────────────────────────────────────
    headers = ['Fecha Ingreso', 'Material', 'Categoría', 'Cantidad', 'Capataz / Receptor', 'Observación']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill_hdr
        c.alignment = center
        c.border = border
    ws.row_dimensions[3].height = 20

    # ── Datos: solo INGRESO_STOCK, todos los registros de la obra ─────────
    ingresos = HistorialMaterial.objects.filter(
        obra=obra,
        tipo_movimiento='INGRESO_STOCK',
    ).select_related('material', 'trabajador_capataz').order_by('fecha_movimiento')

    row = 4
    for i, mov in enumerate(ingresos):
        relleno = fill_alt if i % 2 == 0 else PatternFill(fill_type=None)
        capataz = mov.trabajador_capataz.nombre_completo if mov.trabajador_capataz else '—'
        data = [
            mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M'),
            mov.material.nombre,
            mov.material.get_categoria_display(),
            mov.cantidad,
            capataz,
            mov.observacion or '—',
        ]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=9)
            c.fill = relleno
            c.border = border
            c.alignment = center if col in (1, 3, 4) else left
        ws.row_dimensions[row].height = 14
        row += 1

    if not ingresos.exists():
        ws.merge_cells('A4:F4')
        c = ws.cell(row=4, column=1, value='Sin materiales ingresados a esta obra')
        c.font = Font(italic=True, color='888888', size=9)
        c.alignment = center

    # ── Total ─────────────────────────────────────────────────────────────
    if ingresos.exists():
        total_row = row
        ws.merge_cells(f'A{total_row}:C{total_row}')
        c = ws.cell(row=total_row, column=1, value=f'TOTAL REGISTROS: {ingresos.count()}')
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        c.alignment = left
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).border = border

    # ── Anchos de columna ─────────────────────────────────────────────────
    for col, w in enumerate([18, 32, 20, 12, 30, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A4'


def _sheet_traslados(ws, obra, fecha_inicio, fecha_fin):
    """Hoja 4: historial de traslados de entrada y salida de la obra en el período."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    ws.title = 'Traslados'

    fill_dark   = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    fill_salida = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_entrada= PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_hdr    = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left        = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = f'HISTORIAL DE TRASLADOS — {obra.nombre}'
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = fill_dark
    c.alignment = center
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}'
    c.font = Font(size=9, italic=True, color='555555')
    c.alignment = center
    ws.row_dimensions[2].height = 15

    headers = ['Fecha', 'Trabajador', 'RUT', 'Dirección', 'Obra Origen', 'Obra Destino', 'Observación']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill_hdr
        c.alignment = center
        c.border = border
    ws.row_dimensions[3].height = 20

    traslados = Traslado.objects.filter(
        Q(obra_origen=obra) | Q(obra_destino=obra),
        fecha_traslado__gte=fecha_inicio,
        fecha_traslado__lte=fecha_fin,
        activo=True,
    ).select_related('trabajador', 'obra_origen', 'obra_destino').order_by('fecha_traslado')

    row = 4
    for t in traslados:
        es_salida = (t.obra_origen_id == obra.pk)
        direccion = 'Salida' if es_salida else 'Entrada'
        fill = fill_salida if es_salida else fill_entrada
        vals = [
            t.fecha_traslado,
            t.trabajador.nombre_completo,
            t.trabajador.rut,
            direccion,
            t.obra_origen.nombre if t.obra_origen else '—',
            t.obra_destino.nombre if t.obra_destino else '—',
            t.observaciones or '—',
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=9)
            c.fill = fill
            c.border = border
            c.alignment = center if col in (1, 3, 4) else left
        ws.row_dimensions[row].height = 14
        row += 1

    if row == 4:
        ws.merge_cells('A4:G4')
        c = ws.cell(row=4, column=1, value='Sin traslados registrados en el período')
        c.font = Font(italic=True, color='888888', size=9)
        c.alignment = center

    for col, w in enumerate([14, 32, 16, 10, 28, 28, 36], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A4'


def _sheet_recontrataciones(ws, obra, fecha_inicio, fecha_fin, contratos):
    """Hoja: historial de re-contrataciones en el período."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    fill_hdr = PatternFill(fill_type='solid', fgColor='7C3AED')
    font_hdr = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.title = 'Recontrataciones'
    ws.merge_cells('A1:H1')
    title = ws.cell(row=1, column=1,
        value=f'RE-CONTRATACIONES — {obra.nombre} — {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}')
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal='center')

    headers = ['N°', 'Trabajador', 'RUT', 'Especialidad',
               'Inicio nuevo contrato', 'Contrato anterior #', 'Estado anterior', 'Finiquito subido?']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = center
        cell.border = thin

    for idx, c in enumerate(contratos, 1):
        ant = c.contrato_anterior
        tiene_finiquito = '-'
        if ant:
            tiene_finiquito = 'Sí' if Documento.objects.filter(
                contrato=ant, tipo_documento__nombre='Finiquito Legalizado', activo=True
            ).exists() else 'No'
        row_data = [
            idx,
            c.trabajador.nombre_completo,
            c.trabajador.rut,
            c.especialidad.nombre,
            c.fecha_inicio.strftime('%d/%m/%Y'),
            ant.pk if ant else '',
            ant.estado if ant else '',
            tiene_finiquito,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 3, column=col, value=val)
            cell.border = thin

    if not list(contratos):
        ws.merge_cells('A4:H4')
        cell = ws.cell(row=4, column=1, value='Sin re-contrataciones en el período')
        cell.font = Font(italic=True, color='888888', size=9)
        cell.alignment = Alignment(horizontal='center')

    for col, w in enumerate([5, 30, 14, 18, 20, 18, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A4'


def _generar_excel_cierre(obra, fecha_inicio, fecha_fin, descripcion=''):
    """
    Genera el Excel de cierre mensual con hoja de remuneraciones y checklist documental.
    Retorna (BytesIO, nombre_archivo).
    """
    try:
        import openpyxl
    except ImportError:
        return None, None

    # Contratos del período
    contratos_periodo = Contrato.objects.filter(
        obra=obra, activo=True,
        fecha_inicio__lte=fecha_fin,
    ).filter(
        Q(fecha_termino_real__gte=fecha_inicio) | Q(fecha_termino_real__isnull=True)
    ).select_related('trabajador', 'especialidad').order_by(
        'trabajador__apellidos', 'trabajador__nombres'
    )

    # Tasas vigentes de remuneraciones
    from ..models import ConfigRemuneraciones, LicenciaMedica
    cfg = ConfigRemuneraciones.vigente()
    tasa_afp = float(cfg.tasa_afp_empleado) / 100 if cfg else 0.1058
    tasa_salud = float(cfg.tasa_salud_empleado) / 100 if cfg else 0.07
    tasa_ces = float(cfg.tasa_cesantia_empleado_plazo_fijo) / 100 if cfg else 0.006
    tasa_ces_indef = float(cfg.tasa_cesantia_empleado_indefinido) / 100 if cfg else 0.006

    # Filas para planilla de remuneraciones
    filas_rem = []
    workers_data = []
    seen_ruts = set()
    for c in contratos_periodo:
        inicio_real = max(c.fecha_inicio, fecha_inicio)
        fin_contrato = c.fecha_termino_real or fecha_fin
        fin_real = min(fin_contrato, fecha_fin)
        dias = (fin_real - inicio_real).days + 1
        if dias <= 0:
            continue
        # Días de licencia dentro del período
        dias_licencia = 0
        licencias_c = LicenciaMedica.objects.filter(
            contrato=c, activo=True,
            fecha_inicio__lte=fin_real,
        ).filter(
            Q(fecha_fin__gte=inicio_real) | Q(fecha_fin__isnull=True)
        )
        for lic in licencias_c:
            lic_ini = max(lic.fecha_inicio, inicio_real)
            lic_fin = min(lic.fecha_fin or fin_real, fin_real)
            d = (lic_fin - lic_ini).days + 1
            if d > 0:
                dias_licencia += d
        dias_licencia = min(dias_licencia, dias)
        dias_trabajados = dias - dias_licencia

        sueldo_base_f = float(c.sueldo_base)
        dias_periodo = (fecha_fin - fecha_inicio).days + 1
        sueldo_prop = round(sueldo_base_f * dias_trabajados / dias_periodo, 0) if dias_periodo else 0

        tasa_c = tasa_ces_indef if c.tipo_contrato == 'Indefinido' else tasa_ces
        desc_afp = round(sueldo_prop * tasa_afp, 0)
        desc_salud = round(sueldo_prop * tasa_salud, 0)
        desc_ces = round(sueldo_prop * tasa_c, 0)
        total_desc = desc_afp + desc_salud + desc_ces
        liquido = sueldo_prop - total_desc

        filas_rem.append({
            'nombre': c.trabajador.nombre_completo,
            'rut': c.trabajador.rut,
            'especialidad': c.especialidad.nombre,
            'tipo_contrato': c.tipo_contrato,
            'dias': dias,
            'dias_licencia': dias_licencia,
            'dias_trabajados': dias_trabajados,
            'sueldo_base': sueldo_base_f,
            'sueldo_proporcional': sueldo_prop,
            'desc_afp': desc_afp,
            'desc_salud': desc_salud,
            'desc_cesantia': desc_ces,
            'total_descuentos': total_desc,
            'liquido': liquido,
        })
        if c.trabajador.rut not in seen_ruts:
            seen_ruts.add(c.trabajador.rut)
            workers_data.append({'trabajador': c.trabajador, 'contrato': c, 'especialidad': c.especialidad.nombre})

    # Tipos de documentos
    tipos_personal = list(TipoDocumento.objects.filter(activo=True, nivel='Trabajador', obligatorio=True).order_by('nombre'))
    tipos_contrato_docs = list(TipoDocumento.objects.filter(activo=True, nivel='Contrato', obligatorio=True).order_by('nombre'))
    tipos_obra_docs = list(TipoDocumento.objects.filter(activo=True, nivel='Obra', obligatorio=True).order_by('nombre'))

    # Checklist por trabajador
    worker_checklist = []
    for item in workers_data:
        rut = item['trabajador'].rut
        personal = {}
        for ch in get_checklist_trabajador(rut):
            personal[ch['tipo'].pk] = ch['estado']
        contractual = {}
        for ch in get_checklist_contrato(item['contrato']):
            contractual[ch['tipo'].pk] = ch['estado']
        worker_checklist.append({
            'trabajador': item['trabajador'],
            'especialidad': item['especialidad'],
            'personal': personal,
            'contractual': contractual,
        })

    # Checklist documentos de obra
    obra_docs = {}
    for t in tipos_obra_docs:
        doc = Documento.objects.filter(obra=obra, tipo_documento=t, activo=True).order_by('-fecha_carga').first()
        if doc is None:
            obra_docs[t.pk] = 'pendiente'
        elif doc.esta_vencido:
            obra_docs[t.pk] = 'vencido'
        else:
            obra_docs[t.pk] = 'ok'

    wb = openpyxl.Workbook()

    # Sheet 1: Remuneraciones
    ws1 = wb.active
    _sheet_remuneraciones(ws1, obra, fecha_inicio, fecha_fin, filas_rem, descripcion)

    # Sheet 2+: Checklist (max 40 trabajadores por hoja)
    MAX_PER_SHEET = 40
    chunks = [worker_checklist[i:i + MAX_PER_SHEET] for i in range(0, len(worker_checklist), MAX_PER_SHEET)] if worker_checklist else [[]]
    total_sheets = len(chunks)
    for sheet_idx, chunk in enumerate(chunks, 1):
        ws = wb.create_sheet()
        _sheet_checklist(ws, obra, fecha_inicio, fecha_fin, chunk,
                         tipos_personal, tipos_contrato_docs, tipos_obra_docs,
                         obra_docs, sheet_idx, total_sheets)

    ws_bodega = wb.create_sheet()
    _sheet_bodega(ws_bodega, obra, fecha_inicio, fecha_fin)

    # Hoja 4: Traslados — solo si hay registros en el período
    hay_traslados = Traslado.objects.filter(
        Q(obra_origen=obra) | Q(obra_destino=obra),
        fecha_traslado__gte=fecha_inicio,
        fecha_traslado__lte=fecha_fin,
        activo=True,
    ).exists()
    if hay_traslados:
        ws_traslados = wb.create_sheet()
        _sheet_traslados(ws_traslados, obra, fecha_inicio, fecha_fin)

    # Hoja 5: Re-contrataciones — solo si hay en el período
    recontrataciones = list(Contrato.objects.filter(
        obra=obra, activo=True, es_recontratacion=True,
        fecha_inicio__gte=fecha_inicio,
        fecha_inicio__lte=fecha_fin,
    ).select_related('trabajador', 'especialidad', 'contrato_anterior'))
    if recontrataciones:
        ws_rec = wb.create_sheet()
        _sheet_recontrataciones(ws_rec, obra, fecha_inicio, fecha_fin, recontrataciones)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_mes = fecha_inicio.strftime('%Y_%m')
    slug_desc = ('_' + re.sub(r'[^\w\-]', '_', descripcion))[:30] if descripcion else ''
    fname = f"cierre_{nombre_mes}_{fecha_inicio.strftime('%d')}-{fecha_fin.strftime('%d%m')}{slug_desc}.xlsx"
    return buf, fname


@login_required
def obra_cierre_mensual(request, pk):
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    if request.method == 'POST':
        fecha_inicio_str = request.POST.get('fecha_inicio', '')
        fecha_fin_str = request.POST.get('fecha_fin', '')
        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except (ValueError, TypeError):
            messages.error(request, 'Fechas inválidas. Indica inicio y fin del período.')
            return redirect(f'/obras/{pk}/?tab=cierres')

        # Validar que fecha_inicio <= fecha_fin
        if fecha_inicio > fecha_fin:
            messages.error(request, 'La fecha de inicio del período no puede ser posterior a la fecha de término.')
            return redirect(f'/obras/{pk}/?tab=cierres')

        mes = fecha_inicio.month
        anio = fecha_inicio.year
        descripcion = request.POST.get('descripcion', '').strip()[:120]

        # Verificar si ya existe cierre para este mes/año en la obra
        from django.db import IntegrityError
        cierre_existente = CierreMensual.objects.filter(obra=obra, mes=mes, anio=anio).first()
        if cierre_existente:
            messages.error(
                request,
                f'Ya existe un cierre registrado para {mes}/{anio} en esta obra '
                f'(registrado el {cierre_existente.fecha_cierre.strftime("%d/%m/%Y %H:%M")} por {cierre_existente.usuario_cierre}). '
                f'Si necesitas corregirlo, elimínalo primero (solo admin).'
            )
            return redirect(f'/obras/{pk}/?tab=cierres')

        cierre = CierreMensual(obra=obra, mes=mes, anio=anio,
                               descripcion=descripcion, usuario_cierre=request.user.username)
        if request.FILES.get('archivo_consolidado'):
            cierre.archivo_consolidado = request.FILES['archivo_consolidado']
        else:
            from django.core.files.base import ContentFile
            buf, fname = _generar_excel_cierre(obra, fecha_inicio, fecha_fin, descripcion)
            if buf and fname:
                cierre.archivo_consolidado = ContentFile(buf.read(), name=fname)
        try:
            cierre.save()
        except IntegrityError:
            messages.error(request, f'Ya existe un cierre para {mes}/{anio} en esta obra. No se puede duplicar.')
            return redirect(f'/obras/{pk}/?tab=cierres')
        messages.success(request, f'Cierre {mes}/{anio} registrado para {obra.nombre}. Planilla Excel generada.')

        # Alerta: trabajadores activos en otras obras durante el período
        contratos_periodo = Contrato.objects.filter(
            obra=obra, activo=True,
            fecha_inicio__lte=fecha_fin,
        ).filter(
            Q(fecha_termino_real__gte=fecha_inicio) | Q(fecha_termino_real__isnull=True)
        ).select_related('trabajador')
        multi_obra = []
        for c in contratos_periodo:
            otras = Contrato.objects.filter(
                trabajador=c.trabajador,
                activo=True,
                estado__in=['Vigente', 'Pendiente de Firma'],
            ).exclude(obra=obra).select_related('obra')
            if otras.exists():
                otras_nombres = ', '.join(o.obra.nombre for o in otras)
                multi_obra.append(f'{c.trabajador.nombre_completo} (también en: {otras_nombres})')
        if multi_obra:
            messages.warning(
                request,
                f'Aviso: los siguientes trabajadores tienen contratos activos en otras obras. '
                f'Este cierre solo incluye los valores de "{obra.nombre}": '
                + ' | '.join(multi_obra)
            )
    return redirect(f'/obras/{pk}/?tab=cierres')


@login_required
def obra_cierre_eliminar(request, pk, cierre_pk):
    """Elimina un cierre mensual (solo admin). Permite corregir errores de fecha."""
    if not request.user.is_superuser:
        messages.error(request, 'Solo el administrador puede eliminar cierres.')
        return redirect(f'/obras/{pk}/?tab=cierres')
    cierre = get_object_or_404(CierreMensual, pk=cierre_pk, obra_id=pk)
    if request.method == 'POST':
        if cierre.archivo_consolidado:
            try:
                if os.path.exists(cierre.archivo_consolidado.path):
                    os.remove(cierre.archivo_consolidado.path)
            except Exception:
                pass
        label = f'{cierre.mes}/{cierre.anio}'
        cierre.delete()
        messages.success(request, f'Cierre {label} eliminado. Ahora puedes registrarlo nuevamente.')
    return redirect(f'/obras/{pk}/?tab=cierres')


@login_required
def obra_planilla_remuneraciones(request, pk):
    """Genera una planilla Excel de remuneraciones para un rango de fechas."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse('openpyxl no disponible.', status=500)

    from datetime import date, timedelta

    obra = get_object_or_404(Obra, pk=pk, activo=True)
    if obra.archivada:
        messages.error(request, 'No se puede generar planilla para una obra archivada.')
        return redirect(f'/obras/{pk}/?tab=cierres')
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')

    if not fecha_inicio_str or not fecha_fin_str:
        messages.error(request, 'Debes indicar fecha de inicio y fecha de término.')
        return redirect(f'/obras/{pk}/?tab=cierres')

    try:
        fecha_inicio = date.fromisoformat(fecha_inicio_str)
        fecha_fin = date.fromisoformat(fecha_fin_str)
    except ValueError:
        messages.error(request, 'Formato de fecha inválido.')
        return redirect(f'/obras/{pk}/?tab=cierres')

    if fecha_fin < fecha_inicio:
        messages.error(request, 'La fecha de término debe ser posterior a la fecha de inicio.')
        return redirect(f'/obras/{pk}/?tab=cierres')

    # Contratos que se solapan con el rango: inicio <= fecha_fin y (termino >= fecha_inicio o sin termino)
    contratos = Contrato.objects.filter(
        obra=obra, activo=True,
        fecha_inicio__lte=fecha_fin,
    ).filter(
        Q(fecha_termino_real__gte=fecha_inicio) |
        Q(fecha_termino_real__isnull=True)
    ).select_related('trabajador', 'especialidad').order_by(
        'trabajador__apellidos', 'trabajador__nombres'
    )

    # Si no hay con fecha_termino_real, también incluir los que tienen fecha_termino_estimada
    # (ya cubiertos porque fecha_termino_real es null → pasan el filtro)

    filas = []
    for c in contratos:
        inicio_real = max(c.fecha_inicio, fecha_inicio)
        fin_contrato = c.fecha_termino_real or fecha_fin
        fin_real = min(fin_contrato, fecha_fin)
        dias = (fin_real - inicio_real).days + 1
        if dias <= 0:
            continue
        filas.append({
            'nombre': c.trabajador.nombre_completo,
            'rut': c.trabajador.rut,
            'especialidad': c.especialidad.nombre,
            'tipo_contrato': c.tipo_contrato,
            'fecha_inicio_contrato': c.fecha_inicio,
            'fecha_fin_contrato': c.fecha_termino_real or c.fecha_termino_estimada,
            'dias': dias,
            'sueldo_base': float(c.sueldo_base),
        })

    # Construir Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Planilla Remuneraciones'

    # Título
    ws.merge_cells('A1:J1')
    titulo = ws['A1']
    titulo.value = f'PLANILLA DE REMUNERACIONES — {obra.nombre}'
    titulo.font = Font(bold=True, size=13, color='FFFFFF')
    titulo.fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    periodo = ws['A2']
    periodo.value = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} al {fecha_fin.strftime("%d/%m/%Y")}  |  Generado el {date.today().strftime("%d/%m/%Y")}'
    periodo.font = Font(size=10, italic=True, color='555555')
    periodo.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # Cabeceras
    headers = [
        'N°', 'Nombre Completo', 'RUT', 'Especialidad',
        'Tipo Contrato', 'Días en Período',
        'Sueldo Base ($)', 'Horas Extras ($)', 'Descuentos ($)', 'Total ($)',
    ]
    header_fill = PatternFill(start_color='374151', end_color='374151', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30

    # Filas
    fill_par = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    fill_impar = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    num_font = Font(size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    for i, f in enumerate(filas, 1):
        row = i + 3
        relleno = fill_par if i % 2 == 0 else fill_impar
        data = [
            i, f['nombre'], f['rut'], f['especialidad'],
            f['tipo_contrato'], f['dias'],
            f['sueldo_base'], '', '', '',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = num_font
            cell.fill = relleno
            cell.border = border
            cell.alignment = center if col in (1, 3, 5, 6) else left
        # Formatear sueldo_base como número
        ws.cell(row=row, column=7).number_format = '#,##0'

    # Fila totales
    total_row = len(filas) + 4
    ws.cell(row=total_row, column=1, value='').border = border
    ws.cell(row=total_row, column=2, value=f'TOTAL TRABAJADORES: {len(filas)}').font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=2).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    ws.cell(row=total_row, column=2).border = border
    for col in range(3, 11):
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        ws.cell(row=total_row, column=col).border = border

    # Anchos de columna
    col_widths = [5, 30, 15, 22, 18, 14, 16, 16, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Congelar cabecera
    ws.freeze_panes = 'A4'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"planilla_{obra.nombre.replace(' ', '_')}_{fecha_inicio_str}_{fecha_fin_str}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def obra_upload_doc(request, pk):
    """Upload an obra-level document directly to an obra."""
    from ..forms import DocumentoForm
    from django.urls import reverse
    obra = get_object_or_404(Obra, pk=pk, activo=True)

    if obra.archivada:
        messages.error(request, f'La obra "{obra.nombre}" está archivada. No se pueden subir más documentos.')
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=carpeta')

    tipo_filter = request.GET.get('tipo', '')

    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES, nivel='Obra')
        if form.is_valid():
            doc = form.save(commit=False)
            doc.obra = obra
            doc.usuario_carga = request.user.username
            doc.save()
            messages.success(request, 'Documento cargado correctamente.')
            return redirect(reverse('obra_detail', args=[pk]) + '?tab=carpeta')
    else:
        initial = {}
        if tipo_filter:
            try:
                tipo_obj = TipoDocumento.objects.get(pk=tipo_filter, nivel='Obra', activo=True)
                initial['tipo_documento'] = tipo_obj
            except TipoDocumento.DoesNotExist:
                pass
        form = DocumentoForm(nivel='Obra', initial=initial)

    tipos_obra = TipoDocumento.objects.filter(nivel='Obra', activo=True).order_by('nombre')
    return render(request, 'obras/upload_doc.html', {
        'form': form,
        'obra': obra,
        'tipos_obra': tipos_obra,
    })


@login_required
def obra_doc_delete(request, doc_pk):
    """Soft-delete an obra-level document."""
    from django.urls import reverse
    doc = get_object_or_404(Documento, pk=doc_pk, activo=True)
    obra_pk = doc.obra_id
    if request.method == 'POST':
        doc.activo = False
        doc.save()
        messages.success(request, 'Documento eliminado.')
    return redirect(reverse('obra_detail', args=[obra_pk]) + '?tab=carpeta')


@login_required
def obra_subir_finiquito(request, pk, contrato_pk):
    """Sube el PDF del finiquito firmado y cambia el contrato a estado Finiquitado."""
    from django.urls import reverse
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    contrato = get_object_or_404(Contrato, pk=contrato_pk, obra=obra, activo=True)

    if obra.archivada:
        messages.error(request, f'La obra "{obra.nombre}" está archivada. No se pueden subir más documentos.')
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=dotacion')

    if contrato.estado != 'Finalizado':
        messages.error(request, f'El contrato de {contrato.trabajador.nombre_completo} no está en estado Finalizado.')
        return redirect(reverse('obra_detail', args=[pk]) + '?tab=carpeta')

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo.')
        else:
            ext = archivo.name.rsplit('.', 1)[-1].lower()
            if ext not in ('pdf', 'jpg', 'jpeg', 'png'):
                messages.error(request, 'Solo se aceptan archivos PDF, JPG, JPEG o PNG.')
            else:
                tipo_doc, _ = TipoDocumento.objects.get_or_create(
                    nombre='Finiquito Legalizado',
                    defaults={'nivel': 'Contrato', 'obligatorio': False, 'dias_validez': None, 'activo': True},
                )
                Documento.objects.create(
                    tipo_documento=tipo_doc,
                    contrato=contrato,
                    obra=obra,
                    trabajador_rut=contrato.trabajador.rut,
                    archivo=archivo,
                    usuario_carga=request.user.username,
                )
                _historial(contrato, 'Finiquitado', request.user.username, 'Finiquito firmado cargado al sistema.')
                contrato.estado = 'Finiquitado'
                contrato.save(update_fields=['estado'])
                messages.success(request, f'Finiquito de {contrato.trabajador.nombre_completo} cargado. Contrato Finiquitado.')
                return redirect(reverse('obra_detail', args=[pk]) + '?tab=carpeta')

    return render(request, 'obras/subir_finiquito.html', {
        'obra': obra,
        'contrato': contrato,
    })


@login_required
def obra_cierre_nomina_ajax(request, pk):
    """AJAX: retorna la nómina calculada para un rango de fechas (preview antes de registrar cierre)."""
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    fecha_inicio_str = request.GET.get('fecha_inicio', '')
    fecha_fin_str = request.GET.get('fecha_fin', '')
    try:
        fecha_inicio = date.fromisoformat(fecha_inicio_str)
        fecha_fin = date.fromisoformat(fecha_fin_str)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Fechas inválidas'}, status=400)

    contratos = Contrato.objects.filter(
        obra=obra, activo=True,
        fecha_inicio__lte=fecha_fin,
    ).filter(
        Q(fecha_termino_real__gte=fecha_inicio) | Q(fecha_termino_real__isnull=True)
    ).select_related('trabajador', 'especialidad').order_by(
        'trabajador__apellidos', 'trabajador__nombres'
    )

    filas = []
    seen_ruts = set()
    for c in contratos:
        inicio_real = max(c.fecha_inicio, fecha_inicio)
        fin_contrato = c.fecha_termino_real or fecha_fin
        fin_real = min(fin_contrato, fecha_fin)
        dias = (fin_real - inicio_real).days + 1
        if dias <= 0:
            continue
        rut = c.trabajador.rut
        otras_obras = list(
            Contrato.objects.filter(trabajador=c.trabajador, estado='Vigente', activo=True)
            .exclude(obra=obra).values_list('obra__nombre', flat=True)
        )
        if rut not in seen_ruts:
            seen_ruts.add(rut)
            filas.append({
                'nombre': c.trabajador.nombre_completo,
                'rut': rut,
                'especialidad': c.especialidad.nombre,
                'dias': dias,
                'sueldo_base': float(c.sueldo_base),
                'otras_obras': otras_obras,
            })

    total_sueldo = sum(f['sueldo_base'] for f in filas)
    promedio_dias = round(sum(f['dias'] for f in filas) / len(filas), 1) if filas else 0

    return JsonResponse({'filas': filas, 'total_sueldo': total_sueldo, 'promedio_dias': promedio_dias})


@login_required
def obra_cierre_dotacion_ajax(request, pk):
    """AJAX: retorna la dotación activa de la obra para mostrar en el preview del cierre mensual."""
    obra = get_object_or_404(Obra, pk=pk, activo=True)
    from django.db.models import Count
    contratos = Contrato.objects.filter(
        obra=obra, estado='Vigente', activo=True
    ).select_related('trabajador', 'especialidad')

    # Detectar trabajadores en otras obras
    dotacion = []
    for c in contratos:
        otras_obras = Contrato.objects.filter(
            trabajador=c.trabajador, estado='Vigente', activo=True
        ).exclude(obra=obra).select_related('obra')
        dotacion.append({
            'nombre': c.trabajador.nombre_completo,
            'rut': c.trabajador.rut,
            'especialidad': c.especialidad.nombre,
            'otras_obras': [oc.obra.nombre for oc in otras_obras],
        })

    return JsonResponse({'dotacion': dotacion, 'total': len(dotacion)})
