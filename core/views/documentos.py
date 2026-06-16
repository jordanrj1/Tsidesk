import io, zipfile, os
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.clickjacking import xframe_options_exempt
from django.http import HttpResponse
from django.db.models import Q
from ..models import (Documento, Trabajador, Obra, Contrato, TipoDocumento,
                      Especialidad, get_checklist_trabajador, get_checklist_contrato)


@login_required
def documentos_pendientes(request):
    from datetime import timedelta
    from django.utils import timezone

    hoy = timezone.now().date()
    ESTADOS_ACTIVOS = ('Vigente', 'En Licencia', 'Reactivado')

    # ── 1 & 2. Contratos vencidos y próximos a vencer ────────────────────
    activos_qs = (
        Contrato.objects.filter(activo=True, estado__in=ESTADOS_ACTIVOS)
        .select_related('trabajador', 'obra', 'especialidad')
    )
    vencidos, proximos = [], []
    limite = hoy + timedelta(days=30)
    for c in activos_qs:
        fecha_eff = c.fecha_extension or c.fecha_termino_estimada
        if not fecha_eff:
            continue
        if fecha_eff < hoy:
            vencidos.append({'contrato': c, 'dias': (hoy - fecha_eff).days, 'fecha': fecha_eff})
        elif fecha_eff <= limite:
            proximos.append({'contrato': c, 'dias': (fecha_eff - hoy).days, 'fecha': fecha_eff})
    vencidos.sort(key=lambda x: x['dias'], reverse=True)
    proximos.sort(key=lambda x: x['dias'])

    # ── 3. Pendientes de firma ────────────────────────────────────────────
    pendientes_firma = list(
        Contrato.objects.filter(activo=True, estado='Pendiente de Firma')
        .select_related('trabajador', 'obra', 'especialidad')
        .order_by('creado_el')
    )

    # ── 4. Finiquitos por subir ───────────────────────────────────────────
    terminados_qs = (
        Contrato.objects.filter(activo=True, estado__in=('Finalizado', 'Finiquitado'))
        .select_related('trabajador', 'obra', 'especialidad')
    )
    terminados_ids = list(terminados_qs.values_list('pk', flat=True))
    con_finiquito = set(
        Documento.objects.filter(
            contrato_id__in=terminados_ids,
            tipo_documento__nombre='Finiquito Legalizado',
            activo=True,
        ).values_list('contrato_id', flat=True)
    )
    sin_finiquito = sorted(
        [c for c in terminados_qs if c.pk not in con_finiquito],
        key=lambda c: c.fecha_termino_real or c.fecha_termino_estimada or hoy,
        reverse=True,
    )

    try:
        tipo_finiquito_id = TipoDocumento.objects.get(nombre='Finiquito Legalizado', activo=True).pk
    except TipoDocumento.DoesNotExist:
        tipo_finiquito_id = None

    # ── 5. Documentos obligatorios faltantes — agrupados por obra ────────
    contratos_vigentes = (
        Contrato.objects.filter(activo=True, estado='Vigente')
        .select_related('trabajador', 'obra', 'especialidad')
        .order_by('obra__nombre', 'trabajador__apellidos')
    )

    obras_map = {}          # obra.pk → {'obra': obra, 'grupos': {rut: {...}}}
    personal_vistos = {}    # rut → set de tipo.pk ya añadidos como personal

    for contrato in contratos_vigentes:
        obra = contrato.obra
        rut = contrato.trabajador.rut

        if obra.pk not in obras_map:
            obras_map[obra.pk] = {'obra': obra, 'grupos': {}}
        if rut not in obras_map[obra.pk]['grupos']:
            obras_map[obra.pk]['grupos'][rut] = {
                'trabajador': contrato.trabajador,
                'docs_personales': [],
                'docs_contrato': [],
            }

        # Docs personales: deduplicar globalmente por trabajador
        if rut not in personal_vistos:
            personal_vistos[rut] = set()
        for item in get_checklist_trabajador(rut):
            if not item['tipo'].obligatorio or item['estado'] not in ('pendiente', 'vencido'):
                continue
            if item['tipo'].pk not in personal_vistos[rut]:
                personal_vistos[rut].add(item['tipo'].pk)
                obras_map[obra.pk]['grupos'][rut]['docs_personales'].append(
                    {'tipo': item['tipo'], 'estado': item['estado'], 'doc': item['doc']}
                )

        # Docs de contrato/obra
        for item in get_checklist_contrato(contrato):
            if not item['tipo'].obligatorio or item['estado'] not in ('pendiente', 'vencido'):
                continue
            obras_map[obra.pk]['grupos'][rut]['docs_contrato'].append({
                'tipo': item['tipo'], 'estado': item['estado'],
                'doc': item['doc'], 'obra': obra, 'contrato': contrato,
            })

    # Convertir a listas separadas por estado de obra
    docs_obras_activas = []
    docs_obras_cerradas = []
    for data in obras_map.values():
        grupos = sorted(
            [g for g in data['grupos'].values() if g['docs_personales'] or g['docs_contrato']],
            key=lambda g: g['trabajador'].apellidos,
        )
        if not grupos:
            continue
        total_obra = sum(len(g['docs_personales']) + len(g['docs_contrato']) for g in grupos)
        entry = {'obra': data['obra'], 'grupos': grupos, 'total': total_obra}
        if data['obra'].estado in ('Activa', 'Pausada'):
            docs_obras_activas.append(entry)
        else:
            docs_obras_cerradas.append(entry)

    docs_obras_activas.sort(key=lambda x: x['obra'].nombre)
    docs_obras_cerradas.sort(key=lambda x: x['obra'].nombre)
    total_docs = sum(x['total'] for x in docs_obras_activas + docs_obras_cerradas)
    total_trabajadores = sum(len(x['grupos']) for x in docs_obras_activas + docs_obras_cerradas)

    # ── 6. Obras en cierre ────────────────────────────────────────────────
    obras_en_cierre = []
    for obra in Obra.objects.filter(estado='Cerrada', archivada=False, activo=True).order_by('nombre'):
        vigentes_n = Contrato.objects.filter(obra=obra, activo=True, estado__in=ESTADOS_ACTIVOS).count()
        terminados_n = Contrato.objects.filter(
            obra=obra, activo=True, estado__in=('Finalizado', 'Finiquitado')
        ).count()
        obras_en_cierre.append({'obra': obra, 'vigentes': vigentes_n, 'terminados': terminados_n})

    total_alertas = (
        len(vencidos) + len(proximos) + len(pendientes_firma)
        + len(sin_finiquito) + total_docs + len(obras_en_cierre)
    )

    return render(request, 'documentos/pendientes.html', {
        'vencidos': vencidos,
        'proximos': proximos,
        'pendientes_firma': pendientes_firma,
        'sin_finiquito': sin_finiquito,
        'tipo_finiquito_id': tipo_finiquito_id,
        'docs_obras_activas': docs_obras_activas,
        'docs_obras_cerradas': docs_obras_cerradas,
        'obras_en_cierre': obras_en_cierre,
        'total_docs': total_docs,
        'total_trabajadores': total_trabajadores,
        'total_alertas': total_alertas,
    })


@login_required
def documento_download(request, pk):
    doc = get_object_or_404(Documento, pk=pk)
    response = HttpResponse(doc.archivo, content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{doc.nombre_archivo}"'
    return response


@xframe_options_exempt
@login_required
def documento_preview_papelera(request, pk):
    """Sirve el archivo inline para previsualización en iframe/embed."""
    import mimetypes
    doc = get_object_or_404(Documento, pk=pk)
    content_type, _ = mimetypes.guess_type(doc.archivo.name)
    content_type = content_type or 'application/octet-stream'
    with doc.archivo.open('rb') as f:
        data = f.read()
    response = HttpResponse(data, content_type=content_type)
    nombre = os.path.basename(doc.archivo.name)
    response['Content-Disposition'] = f'inline; filename="{nombre}"'
    return response


@login_required
def documento_delete(request, pk):
    doc = get_object_or_404(Documento, pk=pk, activo=True)
    next_url = request.POST.get('next') or request.GET.get('next') or ''
    if request.method == 'POST':
        doc.activo = False
        doc.save()
        from django.contrib import messages
        messages.success(request, f'Documento "{doc.tipo_documento.nombre}" movido a la papelera.')
        # Sincronizar estado del contrato según el documento eliminado
        if doc.contrato:
            nombre_tipo = doc.tipo_documento.nombre
            # Borrar contrato firmado → revertir Vigente a Pendiente de Firma
            if nombre_tipo == 'Contrato de Trabajo Firmado' and doc.contrato.estado == 'Vigente':
                queda_firmado = Documento.objects.filter(
                    tipo_documento__nombre='Contrato de Trabajo Firmado',
                    contrato=doc.contrato, activo=True
                ).exists()
                if not queda_firmado:
                    doc.contrato.estado = 'Pendiente de Firma'
                    doc.contrato.save(update_fields=['estado'])
            # Borrar Anexo firmado → limpiar extensión si no quedan otros
            elif nombre_tipo == 'Anexo de Contrato':
                queda_anexo = Documento.objects.filter(
                    tipo_documento__nombre='Anexo de Contrato',
                    contrato=doc.contrato, activo=True
                ).exists()
                if not queda_anexo:
                    doc.contrato.fecha_extension = None
                    doc.contrato.save(update_fields=['fecha_extension'])
        return redirect(next_url) if next_url else redirect('documentos_central')
    return redirect(next_url) if next_url else redirect('documentos_central')


@login_required
def documento_upload_rapido(request):
    """Upload de un documento específico desde cualquier pantalla (modal), redirige a 'next'."""
    from django.contrib import messages
    from ..models import TipoDocumento, Trabajador, Contrato
    from ..forms import DocumentoForm

    next_url = request.POST.get('next') or request.GET.get('next') or 'documentos_pendientes'

    if request.method == 'POST':
        trabajador_rut = request.POST.get('trabajador_rut', '').strip()
        tipo_id = request.POST.get('tipo_documento_id')
        contrato_id = request.POST.get('contrato_id')
        archivo = request.FILES.get('archivo')
        fecha_vencimiento = request.POST.get('fecha_vencimiento') or None

        try:
            tipo = TipoDocumento.objects.get(pk=tipo_id, activo=True)
        except TipoDocumento.DoesNotExist:
            messages.error(request, 'Tipo de documento no válido.')
            return redirect(next_url)

        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo.')
            return redirect(next_url)

        doc = Documento(
            tipo_documento=tipo,
            trabajador_rut=trabajador_rut if tipo.nivel == 'Trabajador' else None,
            usuario_carga=request.user.username,
            archivo=archivo,
        )
        if fecha_vencimiento:
            from datetime import date
            try:
                doc.fecha_vencimiento = date.fromisoformat(fecha_vencimiento)
            except ValueError:
                pass

        if contrato_id and tipo.nivel == 'Contrato':
            try:
                contrato_obj = Contrato.objects.get(pk=contrato_id)
                doc.contrato = contrato_obj
                doc.trabajador_rut = contrato_obj.trabajador.rut
                doc.trabajador_nombre = contrato_obj.trabajador.nombre_completo
            except Contrato.DoesNotExist:
                pass

        if tipo.nivel == 'Trabajador' and trabajador_rut:
            try:
                trab = Trabajador.objects.get(rut=trabajador_rut)
                doc.trabajador_nombre = trab.nombre_completo
            except Trabajador.DoesNotExist:
                pass

        doc.save()
        messages.success(request, f'"{tipo.nombre}" cargado correctamente.')

    return redirect(next_url)


@login_required
def papelera_documentos(request):
    from django.contrib import messages as msg_module
    qs = Documento.objects.filter(activo=False).select_related(
        'tipo_documento', 'obra', 'contrato', 'contrato__trabajador', 'contrato__obra'
    ).order_by('-fecha_carga')

    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(
            Q(trabajador_rut__icontains=q) |
            Q(trabajador_nombre__icontains=q) |
            Q(contrato__trabajador__nombres__icontains=q) |
            Q(contrato__trabajador__apellidos__icontains=q) |
            Q(tipo_documento__nombre__icontains=q)
        ).distinct()

    return render(request, 'documentos/papelera.html', {'documentos': qs, 'q': q})


@login_required
def documento_restore(request, pk):
    from django.contrib import messages
    doc = get_object_or_404(Documento, pk=pk, activo=False)
    next_url = request.POST.get('next') or request.GET.get('next') or ''

    if request.method == 'POST':
        # Si ya existe un doc activo del mismo tipo para el mismo contrato/trabajador,
        # lo movemos a papelera antes de restaurar (revertir a versión anterior).
        desplazados = []
        if doc.contrato_id:
            conflictos = Documento.objects.filter(
                contrato_id=doc.contrato_id,
                tipo_documento_id=doc.tipo_documento_id,
                activo=True,
            ).exclude(pk=pk)
        elif doc.trabajador_rut:
            conflictos = Documento.objects.filter(
                trabajador_rut=doc.trabajador_rut,
                contrato__isnull=True,
                tipo_documento_id=doc.tipo_documento_id,
                activo=True,
            ).exclude(pk=pk)
        else:
            conflictos = Documento.objects.none()

        for c in conflictos:
            c.activo = False
            c.save(update_fields=['activo'])
            desplazados.append(c)

        doc.activo = True
        doc.save(update_fields=['activo'])

        if desplazados:
            messages.success(
                request,
                f'"{doc.tipo_documento.nombre}" restaurado. '
                f'La versión anterior ({desplazados[0].fecha_carga.strftime("%d/%m/%Y")}) '
                f'fue movida a papelera.'
            )
        else:
            messages.success(request, f'"{doc.tipo_documento.nombre}" restaurado correctamente.')

    return redirect(next_url) if next_url else redirect('papelera_documentos')


@login_required
def documento_hard_delete(request, pk):
    from django.contrib import messages
    doc = get_object_or_404(Documento, pk=pk, activo=False)
    if request.method == 'POST':
        nombre_tipo = doc.tipo_documento.nombre
        if doc.archivo:
            try:
                if os.path.exists(doc.archivo.path):
                    os.remove(doc.archivo.path)
            except Exception:
                pass
        doc.delete()
        messages.success(request, f'Documento "{nombre_tipo}" eliminado permanentemente.')
    return redirect('papelera_documentos')


@login_required
def documentos_batch_download(request):
    """Descarga selectiva: recibe lista de IDs y genera un ZIP con esos documentos."""
    if request.method == 'POST':
        pks = request.POST.getlist('doc_ids')
    else:
        pks = request.GET.getlist('doc_ids')

    if not pks:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'No se seleccionaron documentos.')
        return redirect('documentos_pendientes')

    docs = Documento.objects.filter(pk__in=pks, activo=True).select_related('tipo_documento', 'contrato')
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for doc in docs:
            if doc.archivo and os.path.exists(doc.archivo.path):
                ref = doc.trabajador_rut or (f"contrato{doc.contrato_id}" if doc.contrato_id else "obra")
                nombre = f"{ref}_{doc.tipo_documento.nombre}_{doc.nombre_archivo}"
                zf.write(doc.archivo.path, nombre)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="documentos_seleccionados.zip"'
    return response


@login_required
def documentos_central(request):
    """Archivo histórico completo: activos + eliminados, agrupados por obra."""
    from django.db.models import Q

    obra_id       = request.GET.get('obra_id', '')
    tipo_id       = request.GET.get('tipo_id', '')
    estado_filter = request.GET.get('estado', '')
    q             = request.GET.get('q', '')
    has_filter    = bool(q or obra_id or tipo_id or estado_filter)

    # Sin filtro activo=True — mostramos todo el historial
    qs = Documento.objects.select_related(
        'tipo_documento', 'obra',
        'contrato', 'contrato__obra',
        'contrato__trabajador', 'contrato__especialidad',
    ).order_by('-fecha_carga')

    if q:
        qs = qs.filter(
            Q(trabajador_rut__icontains=q) |
            Q(trabajador_nombre__icontains=q) |
            Q(contrato__trabajador__nombres__icontains=q) |
            Q(contrato__trabajador__apellidos__icontains=q) |
            Q(tipo_documento__nombre__icontains=q)
        ).distinct()
    if tipo_id:
        qs = qs.filter(tipo_documento_id=tipo_id)
    if obra_id:
        qs = qs.filter(Q(obra_id=obra_id) | Q(contrato__obra_id=obra_id))

    total_qs = qs.count()
    docs_list = list(qs[:500])
    truncado = total_qs > 500

    if estado_filter:
        docs_list = [d for d in docs_list if d.estado_visual == estado_filter]

    # Enriquecer docs personales con objeto Trabajador (si aún existe)
    ruts_personal = {d.trabajador_rut for d in docs_list if d.trabajador_rut and not d.contrato}
    if ruts_personal:
        from ..models import Trabajador as TrabajadorModel
        trab_map = {t.rut: t for t in TrabajadorModel.objects.filter(rut__in=ruts_personal)}
    else:
        trab_map = {}
    for doc in docs_list:
        doc.trab_personal = trab_map.get(doc.trabajador_rut) if (doc.trabajador_rut and not doc.contrato) else None

    # Agrupar en 3 buckets por estado de obra
    _activas    = {}
    _en_cierre  = {}
    _archivadas = {}
    sin_obra_docs = []

    for doc in docs_list:
        obra_obj = doc.obra or (doc.contrato.obra if doc.contrato else None)
        if obra_obj:
            if obra_obj.archivada:
                bucket = _archivadas
            elif obra_obj.estado == 'Cerrada':
                bucket = _en_cierre
            else:
                bucket = _activas
            if obra_obj.pk not in bucket:
                bucket[obra_obj.pk] = {'obra': obra_obj, 'docs': []}
            bucket[obra_obj.pk]['docs'].append(doc)
        else:
            sin_obra_docs.append(doc)

    grupos_activas    = sorted(_activas.values(),    key=lambda x: x['obra'].nombre)
    grupos_en_cierre  = sorted(_en_cierre.values(),  key=lambda x: x['obra'].nombre)
    grupos_archivadas = sorted(_archivadas.values(), key=lambda x: x['obra'].nombre)
    total = len(docs_list)  # puede ser menor que total_qs si estado_filter redujo la lista

    from ..models import Obra, TipoDocumento
    return render(request, 'documentos/central.html', {
        'has_filter':        has_filter,
        'grupos_activas':    grupos_activas,
        'grupos_en_cierre':  grupos_en_cierre,
        'grupos_archivadas': grupos_archivadas,
        'sin_obra_docs':     sin_obra_docs,
        'obras':  Obra.objects.order_by('nombre'),
        'tipos':  TipoDocumento.objects.filter(activo=True).order_by('nombre'),
        'obra_id':       obra_id,
        'tipo_id':       tipo_id,
        'estado_filter': estado_filter,
        'q':       q,
        'total':   total,
        'truncado': truncado,
    })


@login_required
def documentos_exportar_excel(request):
    """Exporta la lista de documentos pendientes/vencidos a Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl no disponible.', status=500)

    obra_id = request.GET.get('obra_id', '')
    especialidad_id = request.GET.get('especialidad_id', '')

    contratos_activos = Contrato.objects.filter(
        estado='Vigente', activo=True
    ).select_related('trabajador', 'obra', 'especialidad')
    if obra_id:
        contratos_activos = contratos_activos.filter(obra_id=obra_id)
    if especialidad_id:
        contratos_activos = contratos_activos.filter(especialidad_id=especialidad_id)

    pendientes = []
    ruts_procesados = set()
    for contrato in contratos_activos:
        rut = contrato.trabajador.rut
        if rut in ruts_procesados:
            continue
        ruts_procesados.add(rut)
        checklist = get_checklist_trabajador(rut)
        for item in checklist:
            if not item['tipo'].obligatorio:
                continue
            if item['estado'] in ('pendiente', 'vencido'):
                pendientes.append({
                    'nombre': contrato.trabajador.nombre_completo,
                    'rut': rut,
                    'obra': contrato.obra.nombre,
                    'especialidad': contrato.especialidad.nombre,
                    'documento': item['tipo'].nombre,
                    'estado': item['estado'].upper(),
                    'vencimiento': item['doc'].fecha_vencimiento.strftime('%d/%m/%Y') if item['doc'] and item['doc'].fecha_vencimiento else '',
                })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Documentos Pendientes'

    headers = ['Trabajador', 'RUT', 'Obra', 'Especialidad', 'Documento', 'Estado', 'Venció']
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    fill_vencido = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_pendiente = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')

    for row, p in enumerate(pendientes, 2):
        ws.cell(row=row, column=1, value=p['nombre'])
        ws.cell(row=row, column=2, value=p['rut'])
        ws.cell(row=row, column=3, value=p['obra'])
        ws.cell(row=row, column=4, value=p['especialidad'])
        ws.cell(row=row, column=5, value=p['documento'])
        ws.cell(row=row, column=6, value=p['estado'])
        ws.cell(row=row, column=7, value=p['vencimiento'])
        fill = fill_vencido if p['estado'] == 'VENCIDO' else fill_pendiente
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="documentos_pendientes.xlsx"'
    return response
