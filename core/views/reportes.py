import io, json
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.http import HttpResponse
from ..models import Trabajador, Obra, Contrato, BodegaObra, Strike


@login_required
def reportes_consolidados(request):
    # Trabajadores multiobra
    ruts_multiples = (
        Contrato.objects.filter(estado='Vigente', activo=True)
        .values('trabajador')
        .annotate(cnt=Count('id'))
        .filter(cnt__gt=1)
    )
    multiobra = []
    for item in ruts_multiples:
        try:
            t = Trabajador.objects.get(rut=item['trabajador'])
            obras_t = Contrato.objects.filter(
                trabajador=t, estado='Vigente', activo=True
            ).select_related('obra')
            multiobra.append({
                'trabajador': t,
                'contratos': obras_t,
                'total': item['cnt'],
            })
        except Trabajador.DoesNotExist:
            pass

    # Dotación por obra (activas)
    obras_activas = Obra.objects.filter(activo=True, estado='Activa')
    dotacion_por_obra = [
        {'obra': o, 'total': o.contratos.filter(estado='Vigente', activo=True).count()}
        for o in obras_activas
    ]

    # Stock crítico global
    stock_critico = BodegaObra.objects.filter(
        obra__activo=True, obra__estado='Activa'
    ).select_related('obra', 'material')
    stock_critico = [b for b in stock_critico if b.bajo_stock]

    # Trabajadores en lista negra
    lista_negra = Trabajador.objects.filter(en_lista_negra=True, activo=True)

    # Strikes recientes
    strikes_recientes = Strike.objects.filter(activo=True).select_related('trabajador').order_by('-creado_el')[:20]

    # Datos para gráficos (JSON)
    labels_dotacion = [d['obra'].nombre for d in dotacion_por_obra]
    data_dotacion = [d['total'] for d in dotacion_por_obra]

    context = {
        'multiobra': multiobra,
        'dotacion_por_obra': dotacion_por_obra,
        'stock_critico': stock_critico,
        'lista_negra': lista_negra,
        'strikes_recientes': strikes_recientes,
        'labels_dotacion_json': json.dumps(labels_dotacion),
        'data_dotacion_json': json.dumps(data_dotacion),
    }
    return render(request, 'reportes/index.html', context)


@login_required
def reportes_exportar_excel(request):
    """Exporta el reporte de dotación y trabajadores multi-obra a Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl no disponible.', status=500)

    wb = openpyxl.Workbook()

    # Hoja 1: Dotación por obra
    ws1 = wb.active
    ws1.title = 'Dotación por Obra'
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(['Obra', 'Constructora', 'Estado', 'Contratos Vigentes'], 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, o in enumerate(Obra.objects.filter(activo=True).order_by('nombre'), 2):
        ws1.cell(row=row, column=1, value=o.nombre)
        ws1.cell(row=row, column=2, value=o.constructora_mandante)
        ws1.cell(row=row, column=3, value=o.estado)
        ws1.cell(row=row, column=4, value=o.contratos.filter(estado='Vigente', activo=True).count())

    # Hoja 2: Trabajadores activos
    ws2 = wb.create_sheet('Trabajadores Activos')
    for col, h in enumerate(['Nombre', 'RUT', 'Teléfono', 'Correo', 'Obra(s) Activa(s)', 'Strikes'], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font

    for row, t in enumerate(Trabajador.objects.filter(activo=True).order_by('apellidos'), 2):
        obras_t = ', '.join(
            c.obra.nombre for c in Contrato.objects.filter(
                trabajador=t, estado='Vigente', activo=True
            ).select_related('obra')
        )
        ws2.cell(row=row, column=1, value=t.nombre_completo)
        ws2.cell(row=row, column=2, value=t.rut)
        ws2.cell(row=row, column=3, value=t.telefono)
        ws2.cell(row=row, column=4, value=t.correo)
        ws2.cell(row=row, column=5, value=obras_t or 'Sin contrato vigente')
        ws2.cell(row=row, column=6, value=t.strikes_activos)
        if t.en_lista_negra:
            fill_negra = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = fill_negra

    # Hoja 3: Lista Negra
    ws3 = wb.create_sheet('Lista Negra')
    for col, h in enumerate(['Nombre', 'RUT', 'Teléfono'], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    for row, t in enumerate(Trabajador.objects.filter(en_lista_negra=True, activo=True), 2):
        ws3.cell(row=row, column=1, value=t.nombre_completo)
        ws3.cell(row=row, column=2, value=t.rut)
        ws3.cell(row=row, column=3, value=t.telefono)

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_operacional.xlsx"'
    return response
